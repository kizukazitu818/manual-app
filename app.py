import streamlit as st
import time
import os
import json
import datetime
import cv2
import re
import numpy as np
import google.generativeai as genai
from io import BytesIO
import base64
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.drawing.image import Image as ExcelImage
from gtts import gTTS
from google.generativeai.types import HarmCategory, HarmBlockThreshold
from streamlit_drawable_canvas import st_canvas
import streamlit_drawable_canvas as canvas_lib

# --- 0. 決定的修正パッチ ---
def fix_canvas_library():
    def custom_image_to_url(image, width, clamp, channels, output_format, image_id):
        try:
            buffered = BytesIO()
            image.save(buffered, format="PNG")
            img_str = base64.b64encode(buffered.getvalue()).decode()
            return f"data:image/png;base64,{img_str}"
        except Exception:
            return ""

    if hasattr(canvas_lib, 'st_image'):
        canvas_lib.st_image.image_to_url = custom_image_to_url

fix_canvas_library()

# --- 1. アプリ全体の基本設定 ---
st.set_page_config(
    page_title="Nano Factory AI",
    page_icon="📜",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ★UIカスタマイズ★
st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=M+PLUS+Rounded+1c:wght@300;400;700&display=swap');

    html, body, [class*="css"] {
        font-family: 'M PLUS Rounded 1c', sans-serif !important;
    }

    [data-testid="stFileUploaderDropzone"] {
        background-color: #E6F3FF;
        border: 2px dashed #007BFF;
        border-radius: 15px;
        padding: 20px;
    }
    
    [data-testid="stSidebar"] {
        background-color: #E6F3FF;
    }
    
    h1 {
        border-bottom: 5px solid #FFD700;
        padding-bottom: 10px;
    }
    </style>
    """, unsafe_allow_html=True)

# --- 2. モデルリスト取得関数 ---
@st.cache_data(ttl=600)
def get_available_models(api_key):
    default_models = ["gemini-1.5-flash", "gemini-2.0-flash-exp"]
    if not api_key: return default_models
    try:
        genai.configure(api_key=api_key)
        models = []
        for m in genai.list_models():
            if 'generateContent' in m.supported_generation_methods:
                name = m.name.replace("models/", "")
                if "deep-research" in name or "ultra" in name:
                    continue
                models.append(name)
        models.sort()
        prioritized = []
        others = []
        for m in models:
            if "flash" in m: prioritized.append(m)
            else: others.append(m)
        return prioritized + others if (prioritized + others) else default_models
    except Exception:
        return default_models

# --- 3. データ処理用ヘルパー関数群 ---
def clean_timestamp(ts_value):
    if ts_value is None: return 0.0
    if isinstance(ts_value, (int, float)): return float(ts_value)
    s = str(ts_value).strip()
    try:
        return float(s)
    except ValueError:
        if ":" in s:
            parts = s.split(":")
            if len(parts) == 2:
                try: return float(parts[0]) * 60 + float(parts[1])
                except: pass
        numbers = re.findall(r"\d+\.?\d*", s)
        if numbers: return float(numbers[0])
    return 0.0

def extract_frame_as_pil(video_path, seconds):
    cap = cv2.VideoCapture(video_path)
    cap.set(cv2.CAP_PROP_POS_MSEC, seconds * 1000)
    ret, frame = cap.read()
    cap.release()
    if ret:
        frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        return PILImage.fromarray(frame)
    return None

@st.cache_data
def generate_audio_bytes(text):
    try:
        if not text: return None
        tts = gTTS(text=text, lang='ja')
        fp = BytesIO()
        tts.write_to_fp(fp)
        fp.seek(0)
        return fp.read()
    except Exception:
        return None

# --- 4. Excel作成関数 ---
def create_excel_file(steps, m_num, m_author, m_date, video_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "作業手順書"

    header_font = Font(bold=True, size=16)
    meta_font = Font(size=11)
    title_font = Font(bold=True, size=12)
    normal_font = Font(size=11)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    ws['A1'] = f"No: {m_num}"
    ws['A1'].font = Font(bold=True, size=11)
    ws['C1'] = f"作成日: {m_date.strftime('%Y/%m/%d')}"
    ws['C1'].font = meta_font
    ws['C1'].alignment = Alignment(horizontal='right')
    ws['C2'] = f"作成者: {m_author}"
    ws['C2'].font = meta_font
    ws['C2'].alignment = Alignment(horizontal='right')
    ws.merge_cells('A3:C3')
    ws['A3'] = "標準作業手順書"
    ws['A3'].font = header_font
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')

    start_row = 5
    ws[f'A{start_row}'] = "No."
    ws[f'B{start_row}'] = "作業画像"
    ws[f'C{start_row}'] = "作業内容・手順"
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 55
    for col in ['A', 'B', 'C']:
        cell = ws[f'{col}{start_row}']
        cell.font = title_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    current_row = start_row + 1
    for i, step in enumerate(steps, 1):
        ws.row_dimensions[current_row].height = 180
        cell_no = ws[f'A{current_row}']
        cell_no.value = i
        cell_no.alignment = Alignment(horizontal='center', vertical='center')
        cell_no.border = thin_border
        
        cell_img = ws[f'B{current_row}']
        cell_img.border = thin_border
        
        # --- 画像合成ロジック ---
        final_img = None
        
        if video_path:
            ts = clean_timestamp(step.get('timestamp', 0))
            if ts >= 0:
                final_img = extract_frame_as_pil(video_path, ts)

        # ★修正：画像データ(edited_image_data)を使って合成
        if final_img and 'edited_image_data' in step and step['edited_image_data'] is not None:
            try:
                drawing_layer = PILImage.fromarray(step['edited_image_data'].astype('uint8'), 'RGBA')
                drawing_layer = drawing_layer.resize(final_img.size, PILImage.Resampling.LANCZOS)
                final_img.paste(drawing_layer, (0, 0), drawing_layer)
            except Exception as e:
                print(f"Image merge error: {e}")

        if final_img:
            try:
                final_img.thumbnail((320, 240))
                img_byte_arr = BytesIO()
                final_img.save(img_byte_arr, format='PNG')
                img_byte_arr.seek(0)
                excel_img = ExcelImage(img_byte_arr)
                excel_img.anchor = f'B{current_row}'
                ws.add_image(excel_img)
            except Exception:
                cell_img.value = "[画像処理エラー]"
        else:
            cell_img.value = "[画像なし]"

        cell_text = ws[f'C{current_row}']
        cell_text.value = f"【{step['title']}】\n\n{step['text']}"
        cell_text.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cell_text.border = thin_border
        cell_text.font = normal_font
        current_row += 1

    output = BytesIO()
    wb.save(output)
    return output.getvalue()

# --- 5. Gemini API処理 ---
def process_video_with_gemini(video_path, api_key, selected_model):
    genai.configure(api_key=api_key)
    progress_bar = st.progress(0, text="準備中...")
    try:
        progress_bar.progress(10, text="📤 動画をAIサーバーにアップロード中...")
        video_file = genai.upload_file(path=video_path)
        
        while video_file.state.name == "PROCESSING":
            progress_bar.progress(30, text="⏳ AI側で動画を処理しています...（数秒〜数分）")
            time.sleep(2)
            video_file = genai.get_file(video_file.name)
            
        if video_file.state.name == "FAILED":
            raise ValueError("動画の処理に失敗しました。")

        progress_bar.progress(60, text=f"🤖 マニュアルを生成中...（モデル: {selected_model}）")
        model = genai.GenerativeModel(model_name=selected_model)
        
        prompt = """
        あなたは製造現場の熟練管理者です。添付の動画を見て、新人作業員のための「標準作業手順書」を作成してください。
        以下のJSON形式で出力してください:
        [
            {"title": "手順の見出し", "text": "具体的な作業内容。", "timestamp": 5.5},...
        ]
        注意点: 
        - timestampは必ず「秒数（数値）」だけにしてください。（例: 5.5）
        """
        safe = [
            {"category": HarmCategory.HARM_CATEGORY_HARASSMENT, "threshold": HarmBlockThreshold.BLOCK_NONE},
            {"category": HarmCategory.HARM_CATEGORY_HATE_SPEECH, "threshold": HarmBlockThreshold.BLOCK_NONE},
            {"category": HarmCategory.HARM_CATEGORY_SEXUALLY_EXPLICIT, "threshold": HarmBlockThreshold.BLOCK_NONE},
            {"category": HarmCategory.HARM_CATEGORY_DANGEROUS_CONTENT, "threshold": HarmBlockThreshold.BLOCK_NONE},
        ]
        response = model.generate_content(
            [video_file, prompt],
            generation_config={"response_mime_type": "application/json"},
            safety_settings=safe
        )
        progress_bar.progress(100, text="完了！")
        time.sleep(1)
        progress_bar.empty()
        return json.loads(response.text)
    except Exception as e:
        if "429" in str(e):
            st.error(f"⚠️ API制限エラー: '{selected_model}' は利用不可または制限超過です。'gemini-1.5-flash' を選んでください。")
        else:
            st.error(f"エラーが発生しました: {e}")
        return []

# --- 6. サーバー掃除機能 ---
def clear_api_storage(api_key):
    if not api_key:
        st.sidebar.error("APIキーを入力してください")
        return
    try:
        genai.configure(api_key=api_key)
        files = list(genai.list_files())
        if not files:
            st.sidebar.success("削除するファイルはありませんでした。")
            return
        count = 0
        progress = st.sidebar.progress(0, text="削除中...")
        for i, f in enumerate(files):
            try: genai.delete_file(f.name)
            except: pass
            count += 1
            progress.progress((i + 1) / len(files))
        progress.empty()
        st.sidebar.success(f"🧹 {count}個のファイルを削除しました！")
    except Exception as e:
        st.sidebar.error(f"削除エラー: {e}")

# --- 7. サイドバー設定 ---
with st.sidebar:
    try: st.image("nano_banana.png", use_container_width=True)
    except:
        st.header("🍌 Nano Banana")

    st.markdown("### Manufacturing AI Tools")
    st.divider()
    st.header("設定")
    api_key = st.text_input("Google API Key", type="password")
    st.divider()
    st.header("🧠 AIモデル選択")
    
    if api_key:
        available_models = get_available_models(api_key)
        scenario = st.radio(
            "どのような視点の手順書を作成しますか？",
            ["🔧 メカニック視点", "🛡️ 安全管理者視点", "📹 解析・記録視点", "🚀 標準"],
            index=3
        )
        recommended_keyword = "gemini-1.5-flash"
        if "メカニック" in scenario: recommended_keyword = "gemini-2.5"
        elif "安全" in scenario: recommended_keyword = "gemini-3"
        elif "解析" in scenario: recommended_keyword = "robotics"
        
        default_index = 0
        found = False
        for i, m in enumerate(available_models):
            if recommended_keyword in m:
                default_index = i
                found = True
                break
        if not found:
            for i, m in enumerate(available_models):
                if "gemini-1.5-flash" in m:
                    default_index = i
                    found = True
                    break
        
        final_model_name = st.selectbox("使用モデル", available_models, index=default_index)
        
        st.divider()
        with st.expander("🛠️ メンテナンス"):
            if st.button("🗑️ サーバーのゴミ箱を空にする", type="secondary"):
                clear_api_storage(api_key)
    else:
        final_model_name = "gemini-1.5-flash"

    st.divider()
    st.header("📄 文書情報")
    manual_number = st.text_input("マニュアル番号", value="SOP-001")
    author_name = st.text_input("作成者", value="管理者")
    create_date = st.date_input("作成日", datetime.date.today())

# --- 8. メインエリア ---
st.title("📜 Nano Factory AI")
st.markdown("""<p style='font-size: 1.3rem; font-weight: bold; color: #555; margin-bottom: 20px;'>動画からマニュアルを自動生成・編集・Excel出力まで一気通貫で行います。</p>""", unsafe_allow_html=True)
st.markdown("""<div style='font-size: 1.3rem; font-weight: bold; margin-bottom: 10px; display: flex; align-items: center;'>📂 作業動画をアップロードしてください</div>""", unsafe_allow_html=True)

# セッション状態の初期化
if "edit_mode" not in st.session_state: st.session_state.edit_mode = "list"
if "manual_steps" not in st.session_state: st.session_state.manual_steps = None
if "last_uploaded_file" not in st.session_state: st.session_state.last_uploaded_file = None

uploaded_file = st.file_uploader("動画アップロード", type=["mp4", "mov"], label_visibility="collapsed")

if uploaded_file:
    # 動画切り替え時のリセット処理
    if st.session_state.last_uploaded_file != uploaded_file.name:
        st.session_state.manual_steps = None
        st.session_state.edit_mode = "list"
        st.session_state.last_uploaded_file = uploaded_file.name
        temp_filename = "temp_video.mp4"
        with open(temp_filename, "wb") as f:
            while True:
                chunk = uploaded_file.read(1024*1024)
                if not chunk: break
                f.write(chunk)
    else:
        temp_filename = "temp_video.mp4"

    # --- 画面表示 ---
    if st.session_state.edit_mode == "list":
        # === モード1：リスト表示 & 秒数調整 ===
        st.subheader("🎥 現場動画（元データ）")
        st.video(temp_filename)
        st.divider()

        if st.button("AI解析を実行する", type="primary"):
            if not api_key:
                st.error("⚠️ APIキーが必要です")
            else:
                with st.spinner("AI解析中..."):
                    steps = process_video_with_gemini(temp_filename, api_key, final_model_name)
                    if steps:
                        st.session_state.manual_steps = steps
                        st.rerun()

        if st.session_state.manual_steps:
            st.subheader("📝 編集 & プレビュー")
            st.info("秒数を調整して、ベストな画像を選んでください。お絵かきは「次へ」ボタンを押してから行います。")
            
            steps = st.session_state.manual_steps
            for i, step in enumerate(steps):
                with st.container():
                    st.markdown(f"#### 手順 {i+1}")
                    c1, c2 = st.columns([1.5, 1])
                    with c1:
                        ts = clean_timestamp(step.get('timestamp', 0))
                        new_ts = st.number_input(f"秒数 #{i+1}", value=ts, step=0.1, format="%.1f", key=f"ts_{i}")
                        img = extract_frame_as_pil(temp_filename, new_ts)
                        if img: st.image(img, use_container_width=True)
                        steps[i]['timestamp'] = new_ts
                    with c2:
                        steps[i]['title'] = st.text_input(f"見出し #{i+1}", step['title'], key=f"ti_{i}")
                        steps[i]['text'] = st.text_area(f"説明 #{i+1}", step['text'], height=150, key=f"tx_{i}")
                    st.divider()
            
            col_btn1, col_btn2 = st.columns(2)
            with col_btn1:
                if st.button("🎨 画像を編集（お絵かき）する", type="primary", use_container_width=True):
                    st.session_state.edit_mode = "draw"
                    st.rerun()
            with col_btn2:
                excel_data = create_excel_file(steps, manual_number, author_name, create_date, temp_filename)
                st.download_button("📥 そのままExcel出力", excel_data, f"{manual_number}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

    elif st.session_state.edit_mode == "draw":
        # === モード2：お絵かき集中モード ===
        st.subheader("🎨 画像編集モード")
        st.info("1枚ずつ選択して、矢印や枠線を描き込んでください。")
        
        steps = st.session_state.manual_steps
        step_options = [f"手順 {i+1}: {s['title']}" for i, s in enumerate(steps)]
        selected_option = st.selectbox("編集する画像を選択:", step_options)
        selected_index = step_options.index(selected_option)
        
        t1, t2, t3 = st.columns([1,1,2])
        with t1: mode = st.selectbox("ツール", ["rect", "circle", "line", "text", "transform"], key="draw_tool")
        with t2: color = st.color_picker("色", "#FF0000", key="draw_color")
        with t3: width = st.slider("太さ", 1, 10, 3, key="draw_width")

        target_step = steps[selected_index]
        ts = clean_timestamp(target_step.get('timestamp', 0))
        bg_img = extract_frame_as_pil(temp_filename, ts)
        
        if bg_img:
            display_img = bg_img.copy()
            display_img.thumbnail((800, 800))
            
            # ★修正点：読み込むのは「描画データ(drawing_state)」にする（画像データではない）
            initial_data_json = target_step.get('drawing_state')
            
            canvas_result = st_canvas(
                fill_color="rgba(255, 165, 0, 0.1)",
                stroke_width=width, stroke_color=color,
                background_image=display_img,
                update_streamlit=True,
                height=400,
                drawing_mode=mode,
                # ★修正点：JSONデータを渡す（なければNone）
                initial_drawing=initial_data_json if initial_data_json else None,
                key=f"canvas_editor_{selected_index}",
                display_toolbar=True,
            )
            
            # ★修正点：保存するときに「画像データ」と「描画データ」を両方保存する
            if canvas_result.image_data is not None:
                steps[selected_index]['edited_image_data'] = canvas_result.image_data
            if canvas_result.json_data is not None:
                steps[selected_index]['drawing_state'] = canvas_result.json_data

        st.divider()
        c1, c2 = st.columns(2)
        with c1:
            if st.button("↩️ リストに戻る"):
                st.session_state.edit_mode = "list"
                st.rerun()
        with c2:
            excel_data = create_excel_file(steps, manual_number, author_name, create_date, temp_filename)
            st.download_button("📥 編集完了！Excelをダウンロード", excel_data, f"{manual_number}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary", use_container_width=True)
